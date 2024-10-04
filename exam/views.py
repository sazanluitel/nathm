import os

from django.shortcuts import render, redirect, get_object_or_404
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.workbook import Workbook

from exam.models import Subject
from ismt.settings import MEDIA_ROOT
from django.views import View
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.urls import reverse
from django.http import JsonResponse, FileResponse

from routine.models import ExamProgramRoutine
from students.models import Student


class ExamView(View):
    def get(self, request):
        return render(request, 'dashboard/exam/exam.html')


class ExamAjaxView(View):
    paginate_by = 10

    def get(self, request):
        draw = int(request.GET.get("draw", 1))
        start = int(request.GET.get("start", 0))
        length = int(request.GET.get("length", self.paginate_by))
        search_value = request.GET.get("search[value]", None)
        page_number = (start // length) + 1

        exams = self.get_queryset(search_value)

        paginator = Paginator(exams, length)
        page_exams = paginator.page(page_number)

        data = []
        for exam in page_exams:
            start_date = exam.start_date.strftime("%d %b")
            end_date = exam.end_date.strftime("%d %b")

            start_time = exam.start_time.strftime("%I:%M %p")
            end_time = exam.end_time.strftime("%I:%M %p")

            data.append([
                exam.title,
                exam.program.name if exam.program else "",
                f"{start_date} to {end_date}",
                f"{start_time} - {end_time}",
                self.get_action(exam)
            ])

        return JsonResponse({
            "draw": draw,
            "recordsTotal": paginator.count,
            "recordsFiltered": paginator.count,
            "data": data,
        }, status=200)

    def get_queryset(self, search_value):
        """
        Returns the queryset of exams. 
        If search_value is provided, it filters the exams based on the search criteria.
        """
        exams = ExamProgramRoutine.objects.all().order_by('-id')
        if search_value:
            exams = exams.filter(
                Q(title__icontains=search_value) |
                Q(program__name__icontains=search_value)
            )
        return exams

    def get_action(self, exam):
        student_list = reverse("exam_urls:studentlist", kwargs={'id': exam.id})
        return f'''
            <div class="btn-group">
                <a href="{student_list}" class="btn btn-success btn-sm">View Students</a>
            </div>
        '''


class StudentsProgramListView(View):
    def get(self, request, *args, **kwargs):
        exam_id = kwargs.pop('id', None)
        exam = get_object_or_404(ExamProgramRoutine, id=exam_id)
        return render(request, 'dashboard/exam/studentlist.html', {'exam': exam})


class StudentsExamTemplateDownloadView(View):
    def get(self, request, *args, **kwargs):
        exam_id = kwargs.pop('id', None)
        exam = get_object_or_404(ExamProgramRoutine, id=exam_id)
        students = Student.objects.filter(program=exam.program)

        wb = Workbook()
        ws = wb.active
        ws.title = "Student Results"

        ws["A1"] = "ID"
        ws["B1"] = "Subject"
        ws["C1"] = "Total Marks"
        ws["D1"] = "Theory Marks"
        ws["E1"] = "Practical Marks"
        ws["F1"] = "Marks Obtained"

        row = 3
        for student in students:
            ws.merge_cells(f"A{row}:F{row}")
            top_left_cell = ws[f"A{row}"]
            top_left_cell.alignment = Alignment(horizontal="center",
                                                vertical="center")
            ws[f"A{row}"] = f"{student.user.get_full_name()} - {student.user.email}"
            row += 1

            subjects = student.get_results(exam)
            if subjects:
                for subject in subjects:
                    ws[f"A{row}"] = subject.id
                    ws[f"B{row}"] = subject.routine.module.name
                    ws[f"C{row}"] = subject.total_marks
                    ws[f"D{row}"] = subject.theory_marks
                    ws[f"E{row}"] = subject.practical_marks
                    ws[f"F{row}"] = subject.marks_obtained
                    row += 1
            row += 1

        directory = os.path.join(MEDIA_ROOT, 'exim')
        os.makedirs(directory, exist_ok=True)
        file_path = os.path.join(directory, f'exam_{exam_id}.xlsx')
        wb.save(file_path)

        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=f'Exam_Results_{exam_id}.xlsx')

    def post(self, request, *args, **kwargs):
        exam_id = kwargs.pop('id', None)
        exam = get_object_or_404(ExamProgramRoutine, id=exam_id)

        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return JsonResponse({"error": "No file uploaded."}, status=400)

        # Load the workbook and extract data
        wb = load_workbook(uploaded_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] is not None:
                subject_id = row[0]
                try:
                    subject = Subject.objects.get(id=subject_id)
                    subject.total_marks = row[2]
                    subject.theory_marks = row[3]
                    subject.practical_marks = row[4]
                    subject.marks_obtained = row[5]
                    subject.save()
                except Subject.DoesNotExist:
                    pass

        messages.success(request, "Results updated successfully.")
        return redirect("exam_urls:studentlist", id=exam.id)


class StudentsProgramAjaxView(View):
    paginate_by = 10

    def get(self, request, id):
        draw = int(request.GET.get("draw", 1))
        start = int(request.GET.get("start", 0))
        length = int(request.GET.get("length", self.paginate_by))
        page_number = (start // length) + 1

        exam = get_object_or_404(ExamProgramRoutine, id=id)
        students = Student.objects.filter(program=exam.program)

        paginator = Paginator(students, length)
        page_students = paginator.page(page_number)

        data = []
        for student in page_students:
            full_name = f"{student.user.first_name} {student.user.last_name}"
            data.append([
                full_name,
                self.get_action(student.id, exam.id)
            ])

        return JsonResponse({
            "draw": draw,
            "recordsTotal": paginator.count,
            "recordsFiltered": paginator.count,
            "data": data,
        }, status=200)

    def get_action(self, student_id, exam_id):
        result_url = reverse("exam_urls:results", kwargs={'exam_id': exam_id, 'student_id': student_id})
        return f'''
            <a href="{result_url}" class="btn btn-primary btn-sm">Update Result</a>
        '''


class ResultView(View):
    def get(self, request, *args, **kwargs):
        exam_id = kwargs.get("exam_id", None)
        student_id = kwargs.get("student_id", None)
        exam = get_object_or_404(ExamProgramRoutine, id=exam_id)
        student = get_object_or_404(Student, id=student_id)

        results = student.get_results(exam)
        return render(request, 'dashboard/exam/result.html', {
            'exam': exam,
            'student': student,
            'results': results
        })

    def post(self, request, *args, **kwargs):
        exam_id = kwargs.get("exam_id", None)
        student_id = kwargs.get("student_id", None)
        exam = get_object_or_404(ExamProgramRoutine, id=exam_id)
        student = get_object_or_404(Student, id=student_id)

        subjects = student.get_results(exam)
        for subject in subjects:
            total_marks = request.POST.get(f"subjects[{subject.id}][total_marks]")
            theory_marks = request.POST.get(f"subjects[{subject.id}][theory_marks]")
            practical_marks = request.POST.get(f"subjects[{subject.id}][practical_marks]")
            marks_obtained = request.POST.get(f"subjects[{subject.id}][marks_obtained]")

            subject.total_marks = total_marks
            subject.theory_marks = theory_marks
            subject.practical_marks = practical_marks
            subject.marks_obtained = marks_obtained
            subject.save()

        messages.success(request, "Result updated successfully.")
        return redirect('exam_urls:results', exam_id=exam_id, student_id=student_id)
